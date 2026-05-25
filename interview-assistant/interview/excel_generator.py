"""
Generate GreenNode Interview Assessment Excel file matching the official template.
"""
import logging
import os
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
CATEGORY_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SCORE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")


def generate_assessment_excel(
    assessment: dict,
    candidate_name: str,
    interviewer: str,
    position: str,
    interview_date: str = None,
    output_dir: str = "output",
) -> str:
    """
    Generate an Excel file matching the GreenNode Interview Assessment template.

    Args:
        assessment: Assessment dict from Claude (with functional_skills, greennode_dna, motivation, etc.)
        candidate_name: Candidate's name.
        interviewer: Interviewer's name.
        position: Position applied for.
        interview_date: Date string (defaults to today).
        output_dir: Directory to save the Excel file.

    Returns:
        Path to the generated Excel file.
    """
    if interview_date is None:
        interview_date = date.today().strftime("%Y-%m-%d")

    os.makedirs(output_dir, exist_ok=True)
    safe_name = candidate_name.replace(" ", "_").replace("/", "-")
    filename = f"GreenNode_Interview_Assessment_{safe_name}_{interview_date}.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Interviewer 1"

    # Column widths
    col_widths = {"B": 14, "C": 22, "D": 30, "E": 40, "F": 16, "G": 24}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Row 1: Title
    ws.merge_cells("C1:G1")
    title_cell = ws["C1"]
    title_cell.value = "INTERVIEW ASSESSMENT"
    title_cell.font = Font(bold=True, size=16, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center")

    # Row 2: Candidate info
    info_pairs = [
        ("B2", "Candidate Name", "C2", candidate_name),
        ("D2", "Interviewer", "E2", interviewer),
        ("F2", "Date", "G2", interview_date),
    ]
    for label_cell, label, val_cell, value in info_pairs:
        ws[label_cell].value = label
        ws[label_cell].font = Font(bold=True)
        ws[val_cell].value = value

    # Row 3: Position
    ws["B3"].value = "Position"
    ws["B3"].font = Font(bold=True)
    ws["C3"].value = position

    # Row 5: Column headers
    headers = {
        "B5": "Category",
        "C5": "Sub-Category",
        "D5": "Criterion",
        "E5": "Interview Question",
        "F5": "Interview Score",
        "G5": "Evidence / Notes",
    }
    for cell_ref, header in headers.items():
        cell = ws[cell_ref]
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGNMENT
        cell.border = THIN_BORDER

    # Rows 6-14: Assessment data
    rows_data = _build_rows(assessment)

    for i, row in enumerate(rows_data):
        row_num = 6 + i
        ws[f"B{row_num}"].value = row.get("category", "")
        ws[f"C{row_num}"].value = row.get("subcategory", "")
        ws[f"D{row_num}"].value = row.get("criterion", "")
        ws[f"E{row_num}"].value = row.get("question", "")
        ws[f"F{row_num}"].value = row.get("score")
        ws[f"G{row_num}"].value = row.get("evidence", "")

        for col in ["B", "C", "D", "E", "F", "G"]:
            cell = ws[f"{col}{row_num}"]
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGNMENT

        # Category fill
        if row.get("category"):
            ws[f"B{row_num}"].fill = CATEGORY_FILL
            ws[f"B{row_num}"].font = Font(bold=True)
        if row.get("subcategory"):
            ws[f"C{row_num}"].fill = CATEGORY_FILL

        # Score styling
        score_cell = ws[f"F{row_num}"]
        score_cell.alignment = CENTER_ALIGNMENT
        if score_cell.value is not None:
            score_cell.fill = SCORE_FILL

    # Merge category cells
    ws.merge_cells("B6:B8")   # Capability (Functional Skills)
    ws.merge_cells("B9:B11")  # Capability (DNA) — merged with above conceptually
    ws.merge_cells("B12:B14") # Motivation

    # Actually, Capability spans rows 6-11
    ws.unmerge_cells("B6:B8")
    ws.unmerge_cells("B9:B11")
    ws.merge_cells("B6:B11")
    ws["B6"].value = "Capability"
    ws["B6"].fill = CATEGORY_FILL
    ws["B6"].font = Font(bold=True)
    ws["B6"].alignment = Alignment(vertical="center", horizontal="center", text_rotation=90)

    ws.merge_cells("B12:B14")
    ws["B12"].value = "Motivation"
    ws["B12"].fill = CATEGORY_FILL
    ws["B12"].font = Font(bold=True)
    ws["B12"].alignment = Alignment(vertical="center", horizontal="center", text_rotation=90)

    # Merge subcategory cells
    ws.merge_cells("C6:C8")   # Functional Skills
    ws["C6"].alignment = Alignment(vertical="center")
    ws.merge_cells("C9:C11")  # GreenNode's DNA
    ws["C9"].alignment = Alignment(vertical="center")

    # Row 16: Total score
    ws.merge_cells("B16:E16")
    ws["B16"].value = "TOTAL SCORE"
    ws["B16"].font = Font(bold=True, size=12)
    ws["B16"].fill = TOTAL_FILL
    ws["B16"].alignment = Alignment(horizontal="right", vertical="center")
    ws["B16"].border = THIN_BORDER
    for col in ["C", "D", "E"]:
        ws[f"{col}16"].fill = TOTAL_FILL
        ws[f"{col}16"].border = THIN_BORDER

    ws["F16"].value = f"=AVERAGE(F6:F14)"
    ws["F16"].font = Font(bold=True, size=12)
    ws["F16"].fill = TOTAL_FILL
    ws["F16"].alignment = CENTER_ALIGNMENT
    ws["F16"].border = THIN_BORDER
    ws["F16"].number_format = "0.0"

    ws["G16"].value = '=IFERROR(IF(F16>=3,"HIRE",IF(F16>=2.5,"CONSIDER","NOT PROCEED")),"")'
    ws["G16"].font = Font(bold=True, size=12, color="1F4E79")
    ws["G16"].fill = TOTAL_FILL
    ws["G16"].alignment = CENTER_ALIGNMENT
    ws["G16"].border = THIN_BORDER

    # Row 18: Summary
    ws.merge_cells("B18:G18")
    ws["B18"].value = "Summary"
    ws["B18"].font = Font(bold=True, size=11)
    ws["B18"].fill = HEADER_FILL
    ws["B18"].font = HEADER_FONT

    ws.merge_cells("B19:G21")
    ws["B19"].value = assessment.get("summary", "")
    ws["B19"].alignment = WRAP_ALIGNMENT
    ws["B19"].border = THIN_BORDER

    # Score Guide sheet
    _add_score_guide_sheet(wb)

    wb.save(filepath)
    logging.info(f"Excel saved: {filepath}")
    return filepath


def _build_rows(assessment: dict) -> list:
    """Build the 9 data rows from assessment dict."""
    rows = []

    # Functional Skills (rows 6-8)
    func_skills = assessment.get("functional_skills", [])
    for i in range(3):
        skill = func_skills[i] if i < len(func_skills) else {}
        rows.append({
            "category": "Capability" if i == 0 else "",
            "subcategory": "Functional Skills" if i == 0 else "",
            "criterion": skill.get("criterion", f"Functional Skill {i+1}"),
            "question": skill.get("question_asked", ""),
            "score": skill.get("score"),
            "evidence": skill.get("evidence", ""),
        })

    # GreenNode's DNA (rows 9-11)
    dna = assessment.get("greennode_dna", [])
    dna_names = ["Collaboration", "Continuous Learning & Improvement", "Customer Centric"]
    for i, name in enumerate(dna_names):
        item = next((d for d in dna if d.get("criterion", "").lower().startswith(name.split()[0].lower())), {})
        if not item and i < len(dna):
            item = dna[i]
        rows.append({
            "category": "",
            "subcategory": "GreenNode's DNA" if i == 0 else "",
            "criterion": name,
            "question": item.get("question_asked", ""),
            "score": item.get("score"),
            "evidence": item.get("evidence", ""),
        })

    # Motivation (rows 12-14)
    motivation = assessment.get("motivation", [])
    mot_names = ["WHO YOU ARE", "HOW YOU THINK", "WHAT YOU COMMIT"]
    for i, name in enumerate(mot_names):
        item = next((m for m in motivation if name.lower() in m.get("criterion", "").lower()), {})
        if not item and i < len(motivation):
            item = motivation[i]
        rows.append({
            "category": "Motivation" if i == 0 else "",
            "subcategory": name,
            "criterion": item.get("criterion", name),
            "question": item.get("question_asked", ""),
            "score": item.get("score"),
            "evidence": item.get("evidence", ""),
        })

    return rows


def _add_score_guide_sheet(wb: Workbook):
    """Add the Score Guide reference sheet."""
    ws = wb.create_sheet("Score Guide")

    ws["B1"].value = "SCORING GUIDE"
    ws["B1"].font = Font(bold=True, size=14, color="1F4E79")

    headers = {"B2": "Score", "C2": "Meaning", "D2": "Behavioral Indicators (examples)"}
    for cell_ref, header in headers.items():
        cell = ws[cell_ref]
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    guide = [
        (5, "Strong", "Leads independently; clear evidence/metrics; anticipates risks; influences stakeholders; repeatable method."),
        (4, "Above Expectations", "Good evidence; minor gaps; handles complexity with some support; solid structure and ownership."),
        (3, "Meets Expectations", "Basic competence; can execute with guidance; evidence limited but credible; moderate complexity."),
        (2, "Below Expectations", "Reactive; shallow evidence; inconsistent method; relies heavily on others; limited complexity exposure."),
        (1, "Weak / Not demonstrated", "No relevant examples; unclear role/impact; misconceptions; cannot articulate approach."),
    ]

    for i, (score, meaning, indicators) in enumerate(guide):
        row = 3 + i
        ws[f"B{row}"].value = score
        ws[f"B{row}"].alignment = CENTER_ALIGNMENT
        ws[f"B{row}"].border = THIN_BORDER
        ws[f"C{row}"].value = meaning
        ws[f"C{row}"].border = THIN_BORDER
        ws[f"D{row}"].value = indicators
        ws[f"D{row}"].border = THIN_BORDER
        ws[f"D{row}"].alignment = WRAP_ALIGNMENT

    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 80


# Allow running as a standalone test
if __name__ == "__main__":
    import logging
    logging.basicConfig(level=logging.INFO)

    sample = {
        "functional_skills": [
            {"criterion": "Python Programming", "question_asked": "Describe your Python experience", "score": 4, "evidence": "5 years of experience with Django and FastAPI"},
            {"criterion": "System Design", "question_asked": "Design a scalable system", "score": 3, "evidence": "Basic understanding of microservices"},
            {"criterion": "Data Engineering", "question_asked": "ETL pipeline experience", "score": 4, "evidence": "Built Spark pipelines at scale"},
        ],
        "greennode_dna": [
            {"criterion": "Collaboration", "question_asked": "Teamwork example", "score": 4, "evidence": "Led cross-team project"},
            {"criterion": "Continuous Learning & Improvement", "question_asked": "Learning habits", "score": 3, "evidence": "Takes online courses"},
            {"criterion": "Customer Centric", "question_asked": "Customer focus", "score": 3, "evidence": "Some awareness"},
        ],
        "motivation": [
            {"criterion": "WHO YOU ARE", "question_asked": "Values question", "score": 4, "evidence": "Strong integrity shown"},
            {"criterion": "HOW YOU THINK", "question_asked": "Problem solving", "score": 3, "evidence": "Structured approach"},
            {"criterion": "WHAT YOU COMMIT", "question_asked": "Career goals", "score": 4, "evidence": "Clear direction"},
        ],
        "total_score": 3.6,
        "recommendation": "HIRE",
        "summary": "Strong technical candidate with good collaboration skills. Shows clear career direction and integrity.",
    }

    path = generate_assessment_excel(
        assessment=sample,
        candidate_name="Test Candidate",
        interviewer="Test Interviewer",
        position="Software Engineer",
    )
    print(f"Generated: {path}")
